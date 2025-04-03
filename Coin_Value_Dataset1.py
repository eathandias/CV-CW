import os
import torch
import torch.utils.data
import torchvision.transforms as T
from torchvision.models.detection import fasterrcnn_resnet50_fpn
from torchvision.models.detection.faster_rcnn import FastRCNNPredictor
from PIL import Image, ImageDraw, ImageFont
import numpy as np
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import cv2  
from collections import Counter

# Build and load the Faster R-CNN model
def get_fasterrcnn_model(num_classes):
    model = fasterrcnn_resnet50_fpn(pretrained=False)
    in_features = model.roi_heads.box_predictor.cls_score.in_features
    model.roi_heads.box_predictor = FastRCNNPredictor(in_features, num_classes)
    return model

# Process a single image 
def process_image(model, input_image_path, output_image_path, threshold=0.5, confidence_value_threshold=0.5):
    # Load and convert image to RGB
    input_image = Image.open(input_image_path).convert("RGB")
    # Convert the image to a tensor
    transform = T.Compose([T.ToTensor()])
    image_tensor = transform(input_image).unsqueeze(0)  # Shape: [1, C, H, W]
    device = next(model.parameters()).device
    image_tensor = image_tensor.to(device)
    
    # Run model inference
    model.eval()
    with torch.no_grad():
        predictions = model(image_tensor)[0]

    # Get bounding boxes, scores, and labels
    boxes = predictions["boxes"].cpu().numpy()    
    scores = predictions["scores"].cpu().numpy()    
    labels = predictions["labels"].cpu().numpy()    

    # Only keep predictions with scores above the threshold
    keep = scores >= threshold
    boxes = boxes[keep]
    scores = scores[keep]
    labels = labels[keep]

    # Convert the original image to a NumPy array to draw on it
    image_np = np.array(input_image)

    # Map labels to class names
    class_names = {
        0: "background",
        1: "1 Pound",
        2: "20 Pence",
        3: "Five Pence",
        4: "One Penny",
        5: "Ten Pence",
        6: "Two Pence",
        7: "Unknown"
    }

    # Define monetary values for known coin types
    coin_values = {
        "1 Pound":    1.00,
        "20 Pence":   0.20,
        "Five Pence": 0.05,
        "One Penny":  0.01,
        "Ten Pence":  0.10,
        "Two Pence":  0.02
    }

    total_value = 0.0
    details_list = []

    # Loop over detections and draw bounding boxes with labels
    for i in range(len(boxes)):
        x1, y1, x2, y2 = map(int, boxes[i])
        # Draw a white rectangle as the bounding box
        image_np[y1:y2, x1:x1+2, :] = 255   # left edge
        image_np[y1:y2, x2-2:x2, :] = 255     # right edge
        image_np[y1:y1+2, x1:x2, :] = 255     # top edge
        image_np[y2-2:y2, x1:x2, :] = 255     # bottom edge

        class_id = labels[i]
        class_name = class_names.get(class_id, "Unknown")
        confidence = scores[i]

        # Add coin value if detection is a known coin and meets the confidence threshold
        if class_name in coin_values and confidence >= confidence_value_threshold:
            total_value += coin_values[class_name]

        details_list.append(f"Box{i+1}: {class_name}, score={confidence:.2f}, coords=({x1},{y1},{x2},{y2})")

    # Draw the text labels on the image
    output_image = Image.fromarray(image_np)
    draw = ImageDraw.Draw(output_image)
    try:
        font = ImageFont.truetype("arial.ttf", 15)
    except:
        font = ImageFont.load_default()
    
    for i in range(len(boxes)):
        class_id = labels[i]
        class_name = class_names.get(class_id, "Unknown")
        confidence = scores[i]
        x1, y1, x2, y2 = map(int, boxes[i])
        text = f"{class_name} ({confidence:.2f})"
        text_position = (x1, max(y1 - 20, 0))
        text_bbox = draw.textbbox(text_position, text, font=font)
        draw.rectangle(text_bbox, fill="white")
        draw.text(text_position, text, fill="black", font=font)

    # Save the annotated image to the output path
    output_image.save(output_image_path)
    details_str = "; ".join(details_list)
    return total_value, details_str

#  Process all images in Dataset1 folder
def process_dataset1(model, input_folder, output_folder, excel_file, threshold, confidence_value_threshold):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Dataset", "Image Name", "Total Value (GBP)", "Detection Details"])
    
    total_values = []  
    dataset_name = "Dataset1"  
    
    for filename in os.listdir(input_folder):
        if filename.lower().endswith((".jpg", ".jpeg", ".png", ".bmp")):
            input_image_path = os.path.join(input_folder, filename)
            output_image_path = os.path.join(output_folder, "output_" + filename)
            if not os.path.exists(input_image_path):
                print(f"Image {input_image_path} not found, skipping.")
                continue
            total_value, details = process_image(model, input_image_path, output_image_path, threshold, confidence_value_threshold)
            total_values.append(total_value)
            ws.append([dataset_name, filename, round(total_value, 2), details])
    
    wb.save(excel_file)
    print(f"Results for {dataset_name} appended to {excel_file}")
    return total_values

#  Plot distribution of total coin values using a bar chart
def plot_value_distribution_hist(values, dataset_name="Dataset"):
    plt.figure(figsize=(8, 6))
    # You can choose how many bins you want:
    plt.hist(values, bins=20, edgecolor='black', color='blue')
    
    plt.xlabel("Total Value (GBP)")
    plt.ylabel("Number of Images")
    plt.title(f"Distribution of Total Coin Values in {dataset_name}")
    plt.tight_layout()
    plt.show()


#  Main function
def main():
    num_classes = 8  
    model_path = "Coin_Classify_Best.pth"  
    
    #paths for Dataset1
    input_folder = r"C:\Users\eatha\Downloads\dataset2_2025ECPb\dataset1"
    output_folder = r"C:\Users\eatha\Downloads\dataset2_2025ECPb\output\dataset1"
    excel_file = "coin_results_dataset1.xlsx"

    # Load the Faster R-CNN model
    model = get_fasterrcnn_model(num_classes)
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model.load_state_dict(torch.load(model_path, map_location=device))
    model.to(device)

    # Process all images in Dataset1 and record results in an Excel file
    total_values = process_dataset1(model, input_folder, output_folder, excel_file, threshold=0.5, confidence_value_threshold=0.5)
    # Plot the distribution of total coin values
    plot_value_distribution_hist(total_values, dataset_name="Dataset1")

if __name__ == "__main__":
    main()
