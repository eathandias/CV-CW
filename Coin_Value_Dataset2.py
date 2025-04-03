import os
import torch
import torch.utils.data
import torchvision.transforms as T
from torchvision.models.detection import fasterrcnn_resnet50_fpn
from torchvision.models.detection.faster_rcnn import FastRCNNPredictor
from PIL import Image, ImageDraw, ImageFont
import numpy as np
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import cv2  
from collections import Counter

#  Build and load Faster R-CNN model
def get_fasterrcnn_model(num_classes):
    model = fasterrcnn_resnet50_fpn(pretrained=False)
    in_features = model.roi_heads.box_predictor.cls_score.in_features
    model.roi_heads.box_predictor = FastRCNNPredictor(in_features, num_classes)
    return model

#  Process a image 
def process_image(model, input_image_path, output_image_path, threshold=0.5, confidence_value_threshold=0.5):
    # Load and preprocess image
    input_image = Image.open(input_image_path).convert("RGB")
    # Convert the image to a tensor
    transform = T.Compose([T.ToTensor()])
    image_tensor = transform(input_image).unsqueeze(0)  # Shape: [1, C, H, W]
    device = next(model.parameters()).device
    image_tensor = image_tensor.to(device)
    
    # Run inference
    model.eval()
    with torch.no_grad():
        predictions = model(image_tensor)[0]

    # Get predictions 
    boxes = predictions["boxes"].cpu().numpy()  # [N, 4]
    scores = predictions["scores"].cpu().numpy()  # [N]
    labels = predictions["labels"].cpu().numpy()  # [N]

    # Keep only predictions with scores above the threshold
    keep_indices = scores >= threshold
    boxes = boxes[keep_indices]
    scores = scores[keep_indices]
    labels = labels[keep_indices]

    # Convert the original image to a numpy array for drawing
    image_np = np.array(input_image)

    # Map numeric labels to coin class names
    class_names = {
        0: "background",
        1: "1 Pound",
        2: "20 Pence",
        3: "Five Pence",
        4: "One Penny",
        5: "Ten Pence",
        6: "Two Pence",
        7: "Unknown"
    }

    # Monetary values for recognized UK coins
    coin_values = {
        "1 Pound": 1.00,
        "20 Pence": 0.20,
        "Five Pence": 0.05,
        "One Penny": 0.01,
        "Ten Pence": 0.10,
        "Two Pence": 0.02
    }

    total_value = 0.0
    details_list = []

    # Loop over each detection and draw bounding boxes with labels
    for i in range(len(boxes)):
        x1, y1, x2, y2 = map(int, boxes[i])
        # Draw a border for the bounding box
        image_np[y1:y2, x1:x1+2, :] = 255  # left edge
        image_np[y1:y2, x2-2:x2, :] = 255    # right edge
        image_np[y1:y1+2, x1:x2, :] = 255    # top edge
        image_np[y2-2:y2, x1:x2, :] = 255    # bottom edge

        class_id = labels[i]
        class_name = class_names.get(class_id, "Unknown")
        confidence = scores[i]

        # Increase the total coin value if the coin is recognized and confidence is high enough
        if class_name in coin_values and confidence >= confidence_value_threshold:
            total_value += coin_values[class_name]

        details_list.append(
            f"Box{i+1}: {class_name}, score={confidence:.2f}, coords=({x1},{y1},{x2},{y2})"
        )

    # Draw text labels on the image
    output_image = Image.fromarray(image_np)
    draw = ImageDraw.Draw(output_image)
    try:
        font = ImageFont.truetype("arial.ttf", 15)
    except:
        font = ImageFont.load_default()

    for i in range(len(boxes)):
        class_id = labels[i]
        class_name = class_names.get(class_id, "Unknown")
        confidence = scores[i]
        x1, y1, x2, y2 = map(int, boxes[i])
        text = f"{class_name} ({confidence:.2f})"
        text_position = (x1, max(y1 - 20, 0))
        text_bbox = draw.textbbox(text_position, text, font=font)
        draw.rectangle(text_bbox, fill="white")
        draw.text(text_position, text, fill="black", font=font)

    # Save the annotated image to output_image_path
    output_image.save(output_image_path)
    details_str = "; ".join(details_list)
    return total_value, details_str

#  Process Dataset 2 folder
def process_dataset2(model, input_folder, output_folder, excel_file, threshold, confidence_value_threshold):
    # Create output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # Open or create an Excel workbook for logging results
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Dataset", "Image Name", "Total Value (GBP)", "Detection Details"])
    
    total_values = []  # To store the total coin value per image
    dataset_name = "Dataset2"
    
    # Process each image file in the input folder
    for filename in os.listdir(input_folder):
        if filename.lower().endswith((".jpg", ".jpeg", ".png", ".bmp")):
            input_image_path = os.path.join(input_folder, filename)
            output_image_path = os.path.join(output_folder, "output_" + filename)
            if not os.path.exists(input_image_path):
                print(f"Image {input_image_path} not found, skipping.")
                continue
            total_value, details = process_image(
                model, input_image_path, output_image_path, threshold, confidence_value_threshold
            )
            total_values.append(total_value)
            ws.append([dataset_name, filename, round(total_value, 2), details])
    
    wb.save(excel_file)
    print(f"Results for {dataset_name} appended to {excel_file}")
    return total_values

#  Plot distribution of total coin values using a bar chart
def plot_value_distribution_bar(values, dataset_name="Dataset"):
    plt.figure(figsize=(8, 6))
    value_counts = Counter(values)
    sorted_items = sorted(value_counts.items(), key=lambda x: x[0])
    x_vals = [item[0] for item in sorted_items]
    y_vals = [item[1] for item in sorted_items]
    plt.bar(x_vals, y_vals, width=0.005, edgecolor='black')
    plt.xlabel("Total Value (GBP)")
    plt.ylabel("Number of Images")
    plt.title(f"Distribution of Total Coin Values in {dataset_name}")
    plt.xticks(x_vals)
    plt.tight_layout()
    plt.show()

#  Main script 
def main():
    num_classes = 8  
    model_path = "Coin_Classify_Best.pth"  
    
    #paths for Dataset 2
    input_folder = r"C:\Users\eatha\Downloads\dataset2_2025ECPb\dataset2"
    output_folder = r"C:\Users\eatha\Downloads\dataset2_2025ECPb\output\dataset2"
    excel_file = "coin_results_dataset2.xlsx"

    # Load the Faster R-CNN model
    model = get_fasterrcnn_model(num_classes)
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model.load_state_dict(torch.load(model_path, map_location=device))
    model.to(device)

    # Process Dataset 2 and record results in Excel
    total_values = process_dataset2(model, input_folder, output_folder, excel_file,
                                    threshold=0.5, confidence_value_threshold=0.5)
    # Plot the distribution of total coin values
    plot_value_distribution_bar(total_values, dataset_name="Dataset2")

if __name__ == "__main__":
    main()
